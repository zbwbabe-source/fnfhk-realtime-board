import argparse
import csv
import json
import re
import sqlite3
from pathlib import Path

from openpyxl import load_workbook


MONTHS = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]


def clean_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def clean_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def detect_plan_year(sheet_name):
    match = re.search(r"(20\d{2})", sheet_name)
    if not match:
        raise ValueError(f"Could not detect plan year from sheet name: {sheet_name}")
    return int(match.group(1))


def month_code(plan_year, month_index):
    return f"{plan_year % 100:02d}{month_index:02d}"


def classify_monthly_row(season_group, season_name, type_name):
    if season_group == "Grand Total":
        return "grand_total"
    if season_group and season_group.endswith(" Total") and not season_name and not type_name:
        return "season_group_total"
    if season_name and season_name.endswith(" Total") and not type_name:
        return "season_total"
    return "detail"


def parse_monthly_detail(ws, plan_year):
    rows = []
    for row_idx in range(8, 55):
        season_group = clean_text(ws.cell(row_idx, 1).value)
        season_name = clean_text(ws.cell(row_idx, 2).value)
        type_name = clean_text(ws.cell(row_idx, 3).value)

        if not any([season_group, season_name, type_name]):
            continue

        row_kind = classify_monthly_row(season_group, season_name, type_name)
        for month_index, month_name in enumerate(MONTHS, start=1):
            start_col = 4 + (month_index - 1) * 3
            rows.append(
                {
                    "sheet_name": ws.title,
                    "row_kind": row_kind,
                    "season_group": season_group,
                    "season_name": season_name,
                    "type_name": type_name,
                    "month_name": month_name,
                    "month_code": month_code(plan_year, month_index),
                    "sold_amt": clean_number(ws.cell(row_idx, start_col).value),
                    "sold_gross": clean_number(ws.cell(row_idx, start_col + 1).value),
                    "soh_price": clean_number(ws.cell(row_idx, start_col + 2).value),
                }
            )
    return rows


def parse_monthly_targets(ws, plan_year):
    rows = []
    for month_index, month_name in enumerate(MONTHS, start=1):
        col = 4 + (month_index - 1) * 3
        rows.append(
            {
                "sheet_name": ws.title,
                "plan_year": plan_year,
                "month_name": month_name,
                "month_code": month_code(plan_year, month_index),
                "target_amt": clean_number(ws.cell(57, col).value),
            }
        )
    return rows


def parse_yearly_summary(ws, summary_year):
    rows = []
    row_range = range(60, 63) if summary_year == 2026 else range(66, 69)
    for row_idx in row_range:
        category = clean_text(ws.cell(row_idx, 3).value)
        if not category:
            continue
        row = {
            "sheet_name": ws.title,
            "summary_year": summary_year,
            "category": category,
            "ttl_sold_amt": clean_number(ws.cell(row_idx, 40).value),
            "amt_vs_ly": clean_number(ws.cell(row_idx, 41).value),
            "ttl_sold_gross": clean_number(ws.cell(row_idx, 42).value),
            "disc_off": clean_number(ws.cell(row_idx, 43).value),
            "gross_vs_ly": clean_number(ws.cell(row_idx, 44).value) if summary_year == 2026 else None,
        }

        for month_index, month_name in enumerate(MONTHS, start=1):
            start_col = 4 + (month_index - 1) * 3
            rows.append(
                {
                    **row,
                    "month_name": month_name,
                    "month_code": month_code(summary_year, month_index),
                    "sold_amt": clean_number(ws.cell(row_idx, start_col).value),
                    "sold_gross": clean_number(ws.cell(row_idx, start_col + 1).value),
                }
            )
    return rows


def parse_yoy_summary(ws, plan_year):
    rows = []
    for row_idx in range(72, 75):
        category = clean_text(ws.cell(row_idx, 3).value)
        if not category:
            continue
        for month_index, month_name in enumerate(MONTHS, start=1):
            col = 4 + (month_index - 1) * 3
            rows.append(
                {
                    "sheet_name": ws.title,
                    "comparison_label": clean_text(ws.cell(71, 3).value),
                    "category": category,
                    "month_name": month_name,
                    "month_code": month_code(plan_year, month_index),
                    "amt_yoy": clean_number(ws.cell(row_idx, col).value),
                }
            )
    return rows


def build_section3_targets(monthly_detail_rows, plan_year):
    def get_period_season_context(month_index):
        if month_index in (9, 10, 11, 12):
            return "FW", plan_year % 100
        if month_index in (1, 2):
            return "FW", (plan_year - 1) % 100
        return "SS", plan_year % 100

    def parse_season_name(season_name):
        if not season_name:
            return None, None
        text = str(season_name).strip().upper()
        if len(text) < 6:
            return None, None
        year_text = text[:4]
        season_text = text[4:]
        if not year_text.isdigit():
            return None, None
        if season_text == "SS":
            return int(year_text) % 100, "SS"
        if season_text == "FW":
            return int(year_text) % 100, "FW"
        return None, None

    category_specs = [
        ("WEAR", "wear", "의류", "Wear"),
        ("ACCESSORY", "accessory", "악세", "Accessory"),
    ]

    monthly_values = {}
    for month_index, month_name in enumerate(MONTHS, start=1):
        target_season_type, current_yy = get_period_season_context(month_index)
        month_rows = [
            row
            for row in monthly_detail_rows
            if row["row_kind"] == "detail"
            and row["season_group"] == "SEASONAL"
            and row["month_name"] == month_name
        ]

        eligible_rows = []
        for row in month_rows:
            season_yy, season_type = parse_season_name(row["season_name"])
            if season_yy is None or season_type is None:
                continue
            if season_type != target_season_type:
                continue
            if season_yy > current_yy - 1:
                continue
            eligible_rows.append(row)

        wear_amt = sum((row["sold_amt"] or 0.0) for row in eligible_rows if row["type_name"] == "WEAR")
        wear_gross = sum((row["sold_gross"] or 0.0) for row in eligible_rows if row["type_name"] == "WEAR")
        accessory_amt = sum((row["sold_amt"] or 0.0) for row in eligible_rows if row["type_name"] == "ACCESSORY")
        accessory_gross = sum((row["sold_gross"] or 0.0) for row in eligible_rows if row["type_name"] == "ACCESSORY")

        monthly_values[month_name] = {
            "WEAR": (wear_amt, wear_gross),
            "ACCESSORY": (accessory_amt, accessory_gross),
            "TOTAL": (wear_amt + accessory_amt, wear_gross + accessory_gross),
        }

    target_rows = []
    season_cumulative_amt_map = {"WEAR": 0.0, "ACCESSORY": 0.0, "TOTAL": 0.0}
    season_cumulative_gross_map = {"WEAR": 0.0, "ACCESSORY": 0.0, "TOTAL": 0.0}
    prev_target_season_type = None

    for month_index, month_name in enumerate(MONTHS, start=1):
        month_key = month_code(plan_year, month_index)
        target_season_type, _ = get_period_season_context(month_index)

        if target_season_type != prev_target_season_type:
            season_cumulative_amt_map = {"WEAR": 0.0, "ACCESSORY": 0.0, "TOTAL": 0.0}
            season_cumulative_gross_map = {"WEAR": 0.0, "ACCESSORY": 0.0, "TOTAL": 0.0}
            prev_target_season_type = target_season_type

        for category, category_key, category_label_ko, category_label_en in category_specs:
            month_amt, month_gross = monthly_values[month_name][category]
            season_cumulative_amt_map[category] += month_amt
            season_cumulative_gross_map[category] += month_gross

            monthly_discount_rate = 1 - (month_amt / month_gross) if month_gross else None
            cumulative_discount_rate = (
                1 - (season_cumulative_amt_map[category] / season_cumulative_gross_map[category])
                if season_cumulative_gross_map[category]
                else None
            )

            target_rows.append(
                {
                    "sheet_name": "2026 PLAN (2)",
                    "plan_year": plan_year,
                    "month_name": month_name,
                    "month_code": month_key,
                    "category": category,
                    "category_key": category_key,
                    "category_label_ko": category_label_ko,
                    "category_label_en": category_label_en,
                    "target_mode": "monthly",
                    "target_sold_amt": month_amt,
                    "target_sold_gross": month_gross,
                    "target_discount_rate": monthly_discount_rate,
                }
            )
            target_rows.append(
                {
                    "sheet_name": "2026 PLAN (2)",
                    "plan_year": plan_year,
                    "month_name": month_name,
                    "month_code": month_key,
                    "category": category,
                    "category_key": category_key,
                    "category_label_ko": category_label_ko,
                    "category_label_en": category_label_en,
                    "target_mode": "cumulative",
                    "target_sold_amt": season_cumulative_amt_map[category],
                    "target_sold_gross": season_cumulative_gross_map[category],
                    "target_discount_rate": cumulative_discount_rate,
                }
            )

        total_amt, total_gross = monthly_values[month_name]["TOTAL"]
        season_cumulative_amt_map["TOTAL"] += total_amt
        season_cumulative_gross_map["TOTAL"] += total_gross
        total_monthly_discount_rate = 1 - (total_amt / total_gross) if total_gross else None
        total_cumulative_discount_rate = (
            1 - (season_cumulative_amt_map["TOTAL"] / season_cumulative_gross_map["TOTAL"])
            if season_cumulative_gross_map["TOTAL"]
            else None
        )

        for target_mode, target_amt, target_gross, target_discount_rate in [
            ("monthly", total_amt, total_gross, total_monthly_discount_rate),
            ("cumulative", season_cumulative_amt_map["TOTAL"], season_cumulative_gross_map["TOTAL"], total_cumulative_discount_rate),
        ]:
            target_rows.append(
                {
                    "sheet_name": "2026 PLAN (2)",
                    "plan_year": plan_year,
                    "month_name": month_name,
                    "month_code": month_key,
                    "category": "TOTAL",
                    "category_key": "all",
                    "category_label_ko": "전체",
                    "category_label_en": "Total",
                    "target_mode": target_mode,
                    "target_sold_amt": target_amt,
                    "target_sold_gross": target_gross,
                    "target_discount_rate": target_discount_rate,
                }
            )

    return target_rows


def build_section3_targets_json(rows):
    payload = {}
    for row in rows:
        payload.setdefault(row["month_code"], {})
        payload[row["month_code"]].setdefault(row["target_mode"], {})
        payload[row["month_code"]][row["target_mode"]][row["category_key"]] = {
            "category": row["category"],
            "category_label_ko": row["category_label_ko"],
            "category_label_en": row["category_label_en"],
            "target_sold_amt": row["target_sold_amt"],
            "target_sold_gross": row["target_sold_gross"],
            "target_discount_rate": row["target_discount_rate"],
        }
    return payload


def write_csv(path, rows, fieldnames):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_sqlite(path, tables):
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        path.unlink()

    conn = sqlite3.connect(path)
    try:
        for table_name, rows in tables.items():
            if not rows:
                continue

            columns = list(rows[0].keys())
            column_types = {}
            for col in columns:
                sample = next((row[col] for row in rows if row[col] is not None), None)
                column_types[col] = "REAL" if isinstance(sample, (int, float)) else "TEXT"

            column_defs = ", ".join(f'"{col}" {column_types[col]}' for col in columns)
            conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')
            conn.execute(f'CREATE TABLE "{table_name}" ({column_defs})')

            placeholders = ", ".join("?" for _ in columns)
            insert_sql = f'INSERT INTO "{table_name}" ({", ".join(f"""\"{col}\"""" for col in columns)}) VALUES ({placeholders})'
            for row in rows:
                values = [row[col] for col in columns]
                conn.execute(insert_sql, values)
        conn.commit()
    finally:
        conn.close()


def main():
    parser = argparse.ArgumentParser(description="Convert HKMC Old Season Target workbook into normalized CSV and SQLite outputs.")
    parser.add_argument(
        "--input",
        default=r"d:\Downloads\HKMC Old Season Target.xlsx",
        help="Path to the source xlsx file.",
    )
    parser.add_argument(
        "--output-dir",
        default="data/old_season_target",
        help="Directory where CSV and SQLite outputs will be written.",
    )
    args = parser.parse_args()

    input_path = Path(args.input)
    output_dir = Path(args.output_dir)

    wb = load_workbook(input_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    plan_year = detect_plan_year(ws.title)

    monthly_detail = parse_monthly_detail(ws, plan_year)
    monthly_targets = parse_monthly_targets(ws, plan_year)
    summary_2026 = parse_yearly_summary(ws, 2026)
    summary_2025 = parse_yearly_summary(ws, 2025)
    yoy_summary = parse_yoy_summary(ws, plan_year)
    section3_targets = build_section3_targets(monthly_detail, plan_year)
    section3_targets_json = build_section3_targets_json(section3_targets)

    tables = {
        "monthly_detail": monthly_detail,
        "monthly_targets": monthly_targets,
        "summary_2026": summary_2026,
        "summary_2025": summary_2025,
        "yoy_summary": yoy_summary,
        "section3_targets": section3_targets,
    }

    write_csv(output_dir / "monthly_detail.csv", monthly_detail, list(monthly_detail[0].keys()))
    write_csv(output_dir / "monthly_targets.csv", monthly_targets, list(monthly_targets[0].keys()))
    write_csv(output_dir / "summary_2026.csv", summary_2026, list(summary_2026[0].keys()))
    write_csv(output_dir / "summary_2025.csv", summary_2025, list(summary_2025[0].keys()))
    write_csv(output_dir / "yoy_summary.csv", yoy_summary, list(yoy_summary[0].keys()))
    write_csv(output_dir / "section3_targets.csv", section3_targets, list(section3_targets[0].keys()))
    write_sqlite(output_dir / "old_season_target.sqlite", tables)
    (output_dir / "section3_targets.json").write_text(
        json.dumps(section3_targets_json, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    manifest = {
        "source_file": str(input_path),
        "sheet_name": ws.title,
        "plan_year": plan_year,
        "month_code_rule": {month: month_code(plan_year, idx) for idx, month in enumerate(MONTHS, start=1)},
        "section3_target_modes": ["monthly", "cumulative"],
        "section3_target_categories": ["wear", "accessory", "all"],
        "tables": {name: len(rows) for name, rows in tables.items()},
    }
    (output_dir / "manifest.json").write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")

    print(json.dumps(manifest, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
